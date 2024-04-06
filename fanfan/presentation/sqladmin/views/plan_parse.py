from pathlib import Path

from dishka import make_async_container
from fastapi import Request, UploadFile
from openpyxl import load_workbook
from openpyxl.cell import Cell
from sqladmin import BaseView, expose
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError

from fanfan import TEMP_DIR
from fanfan.infrastructure.db import UnitOfWork
from fanfan.infrastructure.db.models import Event, Nomination, Participant
from fanfan.infrastructure.di.config import ConfigProvider
from fanfan.infrastructure.di.db import DbProvider


async def proceed_plan(path: Path, uow: UnitOfWork):
    # Drop schedule first
    await uow.session.execute(text("TRUNCATE TABLE schedule CASCADE"))
    await uow.session.execute(
        text("ALTER SEQUENCE subscriptions_id_seq RESTART WITH 1"),
    )
    await uow.session.execute(text("ALTER SEQUENCE schedule_id_seq RESTART WITH 1"))
    await uow.session.execute(
        text("ALTER SEQUENCE schedule_position_seq RESTART WITH 1"),
    )
    #
    wb = load_workbook(path)
    ws = wb.active
    for row in ws.iter_rows(min_row=3, min_col=2):
        cc: Cell = row[0]
        nc = ws.cell(cc.row + 1, cc.column)
        if cc.font.b:  # Either Event or Nomination
            if nc.font.b or nc.value is None:  # Event
                new_event = Event(title=cc.value)
                uow.session.add(new_event)
            else:  # Nomination
                try:
                    async with uow.session.begin_nested():
                        new_nomination = Nomination(
                            id=nc.value.split()[0],
                            title=cc.value,
                        )
                        uow.session.add(new_nomination)
                except IntegrityError:
                    pass
        else:  # Participant
            participant_title = cc.value.split(", ", 1)[1]
            nomination_id = cc.value.split()[0]
            participant = await uow.participants.get_participant_by_title(
                participant_title,
            )
            if not participant:
                participant = Participant(
                    title=participant_title,
                    nomination_id=nomination_id,
                )
            new_event = Event(title=participant_title, participant=participant)
            uow.session.add(new_event)


class PlanParseView(BaseView):
    name = "Импорт расписания (C2)"
    icon = "fa-solid fa-file-import"

    @expose("/plan_parse", methods=["GET", "POST"])
    async def plan_parse(self, request: Request):
        if request.method == "GET":
            return await self.templates.TemplateResponse(request, "plan_parse.html")
        form = await request.form()
        file: UploadFile = form["file"]
        await file.seek(0)
        content = await file.read()
        path = TEMP_DIR.joinpath(file.filename)
        with open(path, "wb") as f:
            f.write(content)
        container = make_async_container(DbProvider(), ConfigProvider())
        async with container() as request_container:
            uow = await request_container.get(UnitOfWork)
            async with uow:
                try:
                    await proceed_plan(path, uow)
                    await uow.commit()
                except Exception as e:
                    await container.close()
                    return await self.templates.TemplateResponse(
                        request,
                        "plan_parse.html",
                        context={"error": repr(e)},
                    )
        await container.close()
        return await self.templates.TemplateResponse(
            request,
            "plan_parse.html",
            context={"success": True},
        )
