import logging
import typing

from fastapi import Request, UploadFile
from openpyxl import load_workbook
from sqladmin import BaseView, expose
from sqlalchemy import delete, select, text
from sqlalchemy.ext.asyncio import AsyncSession

from fanfan.adapters.db.models import Event, Participant
from fanfan.adapters.db.models.block import Block

if typing.TYPE_CHECKING:
    from dishka import AsyncContainer

logger = logging.getLogger(__name__)


async def proceed_plan(file: typing.BinaryIO, session: AsyncSession) -> None:
    # Constraints must be deferred, so we can swap ids and order
    await session.execute(text("SET CONSTRAINTS ALL DEFERRED"))
    # Delete orphaned events later
    events_to_delete = list(await session.scalars(select(Event)))
    # Delete all blocks
    await session.execute(delete(Block))
    await session.flush()
    # Get everything ready
    wb = load_workbook(file)
    ws = wb.worksheets[0]
    order = 1.0
    for row in ws.iter_rows(min_row=4):
        if row[0].value and row[2].value:  # Block
            title = str(row[2].value)
            block = Block(title=title, start_order=order)
            session.add(block)
            await session.flush([block])
            logger.info("New block %s added", title)
        elif row[1].value:  # Event
            data = str(row[2].value)
            id_ = int(row[1].value)
            event_title = data.split(", ", 1)[1]
            # Try to find a participant
            participant = await session.scalar(
                select(Participant).where(Participant.title.ilike(f"%{event_title}%"))
            )
            if participant:
                logger.info("Found a linked participant for %s", event_title)
            else:
                logger.info(
                    "Orphaned event, make sure event title %s"
                    "matches participant title (if it exist)",
                    event_title,
                )
            event = await session.merge(
                Event(
                    id=id_,
                    title=event_title,
                    order=order,
                    participant_id=participant.id if participant else None,
                )
            )
            for e in events_to_delete:
                if e.id == event.id:
                    events_to_delete.remove(e)
        order += 1.0
    for e in events_to_delete:
        await session.delete(e)


class PlanParseView(BaseView):
    name = "Импорт расписания (C2)"
    icon = "fa-solid fa-file-import"

    @expose("/plan_parse", methods=["GET", "POST"])
    async def plan_parse(self, request: Request):
        if request.method == "GET":
            return await self.templates.TemplateResponse(request, "plan_parse.html")
        form = await request.form()
        file: UploadFile = form["file"]
        container: AsyncContainer = request.state.dishka_container
        session = await container.get(AsyncSession)
        async with session:
            try:
                await proceed_plan(file.file, session)
                await session.commit()
            except Exception as e:
                await container.close()
                logger.exception("Error when parsing plan", exc_info=e)
                return await self.templates.TemplateResponse(
                    request,
                    "plan_parse.html",
                    context={"error": repr(e)},
                )
        return await self.templates.TemplateResponse(
            request,
            "plan_parse.html",
            context={"success": True},
        )
