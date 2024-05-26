import logging
from pathlib import Path

from dishka import make_async_container
from fastapi import Request, UploadFile
from openpyxl import load_workbook
from sqladmin import BaseView, expose

from fanfan.application.dto.event import CreateEventDTO, UpdateEventDTO
from fanfan.application.dto.nomination import CreateNominationDTO
from fanfan.application.dto.participant import CreateParticipantDTO
from fanfan.common import TEMP_DIR
from fanfan.infrastructure.db import UnitOfWork
from fanfan.infrastructure.di.config import ConfigProvider
from fanfan.infrastructure.di.db import DbProvider

logger = logging.getLogger(__name__)


async def proceed_plan(path: Path, uow: UnitOfWork):
    events_to_delete = list(await uow.events.get_all_events())
    order = 1.0
    wb = load_workbook(path)
    ws = wb.worksheets[0]
    for row in ws.iter_rows():
        id_ = row[1]
        data = row[2]
        if id_.value:
            # Nomination
            nomination_id = data.value.split()[0]
            nomination = await uow.nominations.get_nomination(nomination_id)
            if not nomination:
                await uow.nominations.add_nomination(
                    CreateNominationDTO(
                        id=nomination_id,
                        title=ws.cell(data.row - 1, data.column).value,
                    )
                )

            # Participant
            title = data.value.split(", ", 1)[1]
            participant = await uow.participants.get_participant_by_title(title)
            if not participant:
                participant = await uow.participants.add_participant(
                    CreateParticipantDTO(title=title, nomination_id=nomination_id)
                )

            # Finally, event
            event = await uow.events.get_event(id_.value)
            if event:
                await uow.events.update_event(
                    UpdateEventDTO(
                        id=id_.value,
                        order=order,
                        title=participant.title,
                        participant_id=participant.id,
                    ),
                )
                events_to_delete.remove(event)
            else:
                await uow.events.add_event(
                    CreateEventDTO(
                        id=id_.value,
                        order=order,
                        title=participant.title,
                        participant_id=participant.id,
                    )
                )
            order += 1.0
    for e in events_to_delete:
        await uow.events.delete_event(e.id)


class PlanParseView(BaseView):
    name = "Импорт расписания (C2)"
    icon = "fa-solid fa-file-import"

    @expose("/plan_parse", methods=["GET", "POST"])
    async def plan_parse(self, request: Request):
        if request.method == "GET":
            return await self.templates.TemplateResponse(request, "plan_parse.html")
        form = await request.form()
        file: UploadFile = form["file"]
        await file.seek(0)
        content = await file.read()
        path = TEMP_DIR.joinpath(file.filename)
        with open(path, "wb") as f:
            f.write(content)
        container = make_async_container(DbProvider(), ConfigProvider())
        async with container() as request_container:
            uow = await request_container.get(UnitOfWork)
            async with uow:
                try:
                    await proceed_plan(path, uow)
                    await uow.commit()
                except Exception as e:
                    await container.close()
                    logger.exception(e)
                    return await self.templates.TemplateResponse(
                        request,
                        "plan_parse.html",
                        context={"error": repr(e)},
                    )
        await container.close()
        return await self.templates.TemplateResponse(
            request,
            "plan_parse.html",
            context={"success": True},
        )
