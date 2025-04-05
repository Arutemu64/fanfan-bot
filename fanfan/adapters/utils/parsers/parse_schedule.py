import logging
import typing

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from fanfan.adapters.db.models import EventORM, ParticipantORM
from fanfan.adapters.db.models.block import BlockORM

logger = logging.getLogger(__name__)


async def parse_schedule(file: typing.BinaryIO, session: AsyncSession) -> None:
    # Delete orphaned events later
    events_to_delete = list(await session.scalars(select(EventORM)))
    # Delete all blocks
    await session.execute(delete(BlockORM))
    await session.flush()
    # Get everything ready
    wb = load_workbook(file)
    ws = wb.worksheets[0]
    order = 1.0
    for row in ws.iter_rows(min_row=2):
        if row[4].value:  # Event
            card = int(row[3].value)  # Card number
            id_ = int(row[4].value)  # Number
            event_title = str(row[5].value)  # Title
            # Try to find a participant
            participant = await session.scalar(
                select(ParticipantORM).where(ParticipantORM.voting_number == card)
            )
            if participant:
                logger.info("Found a linked participant for %s", event_title)
            else:
                logger.warning(
                    "Orphaned event, as participant with "
                    "voting number %s was not found",
                    card,
                )
            event = await session.merge(
                EventORM(
                    id=id_,
                    title=event_title,
                    order=order,
                    participant_id=participant.id if participant else None,
                )
            )
            for e in events_to_delete:
                if e.id == event.id:
                    events_to_delete.remove(e)
        elif row[0].value and row[0].font.italic:  # Block
            title = str(row[0].value)
            block = BlockORM(title=title, start_order=order)
            session.add(block)
            await session.flush([block])
            logger.info("New block %s added", title)
        order += 1.0
    for e in events_to_delete:
        await session.delete(e)
    # Commit
    await session.commit()
