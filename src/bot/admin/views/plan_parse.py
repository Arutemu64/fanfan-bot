from pathlib import Path

from fastapi import Request, UploadFile
from openpyxl import load_workbook
from openpyxl.cell import Cell
from sqladmin import BaseView, expose
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from src.db import Database
from src.db.database import create_session_pool
from src.db.models import Event, Nomination, Participant


async def drop_schedule(db: Database):
    print("Dropping schedule")
    await db.session.execute(text("TRUNCATE TABLE schedule CASCADE"))
    await db.session.execute(text("ALTER SEQUENCE subscriptions_id_seq RESTART WITH 1"))
    await db.session.execute(text("ALTER SEQUENCE schedule_id_seq RESTART WITH 1"))
    await db.session.execute(
        text("ALTER SEQUENCE schedule_position_seq RESTART WITH 1")
    )
    await db.session.commit()


async def proceed_plan(path: Path, session: AsyncSession):
    db = Database(session)
    await drop_schedule(db)
    wb = load_workbook(path)
    ws = wb.active
    new_nominations = []
    new_participants = []
    new_events = []
    for row in ws.iter_rows(min_row=3, min_col=2):
        cc: Cell = row[0]
        nc = ws.cell(cc.row + 1, cc.column)
        if cc.font.b:  # Either Event or Nomination
            if nc.font.b or nc.value is None:  # Event
                title = cc.value
                event = await db.event.get_by_title(title)
                if not event:
                    new_event = Event(title=title)
                    db.session.add(new_event)
                    await db.session.flush([new_event])
                    print(f"Added new event {new_event.title}")
                    new_events.append(new_event)
            else:  # Nomination
                code = nc.value.split()[0]
                nomination = await db.nomination.get(code)
                if not nomination:
                    new_nomination = Nomination(id=code, title=cc.value)
                    db.session.add(new_nomination)
                    await db.session.flush([new_nomination])
                    print(f"Added new nomination {new_nomination.id}")
                    new_nominations.append(new_nomination)
        else:  # Participant
            title = cc.value.split(", ", 1)[1]
            code = cc.value.split()[0]
            participant = await db.participant.get_by_title(title)
            nomination = await db.nomination.get(code)
            if not participant:
                participant = Participant(title=title, nomination=nomination)
                db.session.add(participant)
                await db.session.flush([participant])
                new_participants.append(participant)
                print(f"Added new participant {participant.title}")
            new_event = Event(title=title)
            participant.event = new_event
            await db.session.flush([new_event, participant])
            new_events.append(new_event)
            print(f"Added new event {participant.event.title}")
    await db.session.commit()
    return {
        "new_nominations_count": len(new_nominations),
        "new_participants_count": len(new_participants),
        "new_events_count": len(new_events),
    }


class PlanParseView(BaseView):
    name = "Импорт расписания (C2)"
    category = "Парсинг"
    icon = "fa-solid fa-file-import"

    @expose("/plan_parse", methods=["GET", "POST"])
    async def plan_parse(self, request: Request):
        if request.method == "GET":
            return await self.templates.TemplateResponse(request, "plan_parse.html")
        form = await request.form()
        file: UploadFile = form["file"]
        await file.seek(0)
        content = await file.read()
        path = Path(__file__).parent.joinpath(file.filename)
        with open(path, "wb") as f:
            f.write(content)
        session_pool = create_session_pool()
        async with session_pool() as session:
            await proceed_plan(path, session)
        return self.templates.TemplateResponse(
            "plan_parse.html", context={"request": request}
        )
