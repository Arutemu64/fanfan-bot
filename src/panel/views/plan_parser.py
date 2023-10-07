from pathlib import Path

import flask
import flask_login
from flask_admin import BaseView, expose
from flask_wtf import FlaskForm
from flask_wtf.file import FileAllowed, FileField, FileRequired
from openpyxl import load_workbook
from openpyxl.cell import Cell
from sqlalchemy import select, text

from src.db.models import Event, Nomination, Participant
from src.panel import db


def drop_schedule():
    print("Dropping schedule")
    db.session.execute(text("TRUNCATE TABLE schedule CASCADE"))
    db.session.execute(text("ALTER SEQUENCE subscriptions_id_seq RESTART WITH 1"))
    db.session.execute(text("ALTER SEQUENCE schedule_id_seq RESTART WITH 1"))
    db.session.execute(text("ALTER SEQUENCE schedule_position_seq RESTART WITH 1"))
    db.session.commit()


def proceed_plan(path: Path):
    drop_schedule()
    wb = load_workbook(path)
    ws = wb.active
    new_nominations = []
    new_participants = []
    new_events = []
    for row in ws.iter_rows(min_row=3, min_col=2):
        cc: Cell = row[0]
        nc = ws.cell(cc.row + 1, cc.column)
        if cc.font.b:  # Either Event or Nomination
            if nc.font.b or nc.value is None:  # Event
                stmt = select(Event).where(Event.title == cc.value).limit(1)
                event = db.session.execute(stmt).scalar_one_or_none()
                if not event:
                    new_event = Event(title=cc.value)
                    db.session.add(new_event)
                    db.session.flush([new_event])
                    print(f"Added new event {new_event.joined_title}")
                    new_events.append(new_event)
            else:  # Nomination
                code = nc.value.split()[0]
                stmt = select(Nomination).where(Nomination.id == code).limit(1)
                nomination = db.session.execute(stmt).scalar_one_or_none()
                if not nomination:
                    new_nomination = Nomination(id=code, title=cc.value)
                    db.session.add(new_nomination)
                    db.session.flush([new_nomination])
                    print(f"Added new nomination {new_nomination.id}")
                    new_nominations.append(new_nomination)
        else:  # Participant
            code = cc.value.split()[0]
            title = cc.value.split(", ", 1)[1]
            stmt = select(Participant).where(Participant.title == title).limit(1)
            participant = db.session.execute(stmt).scalar_one_or_none()
            if not participant:
                participant = Participant(title=title, nomination_id=code)
                db.session.add(participant)
                db.session.flush([participant])
                print(f"Added new participant {participant.title}")
                new_participants.append(participant)
            new_event = Event(participant_id=participant.id)
            db.session.add(new_event)
            db.session.flush([new_event])
            print(f"Added new event {new_event.joined_title}")
            new_events.append(new_event)
    db.session.commit()
    return {
        "new_nominations_count": len(new_nominations),
        "new_participants_count": len(new_participants),
        "new_events_count": len(new_events),
    }


class PlanParserForm(FlaskForm):
    csrf_enabled = False
    file = FileField(validators=[FileRequired(), FileAllowed(["xlsx"], "xlsx only!")])


class PlanParserView(BaseView):
    def is_accessible(self):
        return flask_login.current_user.is_authenticated

    @expose("/", methods=["GET", "POST"])
    def parse_c2(self):
        form = PlanParserForm()
        if form.validate_on_submit():
            f = form.file.data
            path = Path(__file__).parent.joinpath(f.filename)
            f.save(path)
            results = proceed_plan(path)
            flask.flash(
                f"Было добавлено {results['new_nominations_count']} новых номинаций, "
                f"{results['new_participants_count']} новых участников, "
                f"{results['new_events_count']} событий",
                category="info",
            )
            return self.render("plan_parser.html", form=form)
        return self.render("plan_parser.html", form=form)
